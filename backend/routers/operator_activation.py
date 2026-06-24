# backend/routers/operator_activation.py
import os
import shutil
from datetime import datetime
from fastapi import APIRouter, Depends, Form, HTTPException, File, UploadFile
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from backend.database import get_db
from backend.models.operator_activation import (
    OperatorActivationRequest,
    ActivationDocument,
    OperatorActivationRemark,
)

router = APIRouter()

UPLOAD_BASE = "uploads/operator_activation"

VALID_DOC_TYPES = [
    "hard_copy_form",
    "aadhaar_photo",
    "pan_card",
    "passbook",
    "nseit_certificate",
    "excel_sheet",
]


def parse_optional_date(value: str | None):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {value}. Use YYYY-MM-DD.")


# ─────────────────────────────────────────────
# DC ROUTES
# ─────────────────────────────────────────────


@router.post("/submit")
def submit_operator_activation(
    dc_id: int = Form(...),
    district_id: str = Form(...),
    role: str = Form(None),
    name_as_per_aadhaar: str = Form(...),
    registrar_code: str = Form(None),
    ea_code: str = Form(None),
    user_code: str = Form(None),
    nseit_certificate_number: str = Form(None),
    operator_mobile: str = Form(...),
    primary_email: str = Form(None),
    operator_aadhaar: str = Form(None),
    operator_pan: str = Form(None),  # 🌟 Captures the key sent from submit_form.html
    nseit_certification_date: str = Form(None),
    nseit_certificate_expiry_date: str = Form(None),
    pincode: str = Form(None),
    hard_copy_form: UploadFile = File(...),
    aadhaar_photo: UploadFile = File(...),
    pan_card: UploadFile = File(...),
    passbook: UploadFile = File(...),
    nseit_certificate: UploadFile = File(...),
    excel_sheet: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # 🌟 FIXED: Parse HTML calendar text strings into explicit datetime primitives
    cert_date = parse_optional_date(nseit_certification_date)
    expiry_date = parse_optional_date(nseit_certificate_expiry_date)

    new_request = OperatorActivationRequest(
        dc_id=dc_id,
        district_id=district_id,
        role=role,
        name_as_per_aadhaar=name_as_per_aadhaar,
        registrar_code=registrar_code,
        ea_code=ea_code,
        user_code=user_code,
        nseit_certificate_number=nseit_certificate_number,
        operator_mobile=operator_mobile,
        primary_email=primary_email,
        operator_aadhaar=operator_aadhaar,
        # 🌟 FIXED: Maps incoming form payload field key onto your correct database table column name 'pan_number'
        pan_number=operator_pan.strip().upper() if operator_pan else None,
        nseit_certification_date=cert_date,
        nseit_certificate_expiry_date=expiry_date,
        pincode=pincode,
        status="sent_to_chips",
    )
    db.add(new_request)
    db.flush()
    new_request.request_no = f"RP-A{new_request.id:04d}"
    db.flush()

    # 2. Save each file to disk and create a document row
    uploaded_files = {
        "hard_copy_form": hard_copy_form,
        "aadhaar_photo": aadhaar_photo,
        "pan_card": pan_card,
        "passbook": passbook,
        "nseit_certificate": nseit_certificate,
        "excel_sheet": excel_sheet,
    }

    folder = f"{UPLOAD_BASE}/{dc_id}/{new_request.id}"
    os.makedirs(folder, exist_ok=True)

    for doc_type, upload in uploaded_files.items():
        ext = os.path.splitext(upload.filename)[-1]
        file_path = f"{folder}/{doc_type}{ext}"

        with open(file_path, "wb") as f:
            shutil.copyfileobj(upload.file, f)

        file_size = os.path.getsize(file_path)

        doc = ActivationDocument(
            request_id=new_request.id,
            doc_type=doc_type,
            file_path=file_path,
            original_filename=upload.filename,
            file_size_bytes=file_size,
            mime_type=upload.content_type,
        )
        db.add(doc)

    db.commit()
    db.refresh(new_request)

    return {
        "status": "success",
        "message": "Operator activation request submitted successfully.",
        "request_id": new_request.id,
        "status": new_request.status,
    }


# backend/routers/operator_activation.py


@router.get("/dc/{dc_id}")
def get_dc_requests(dc_id: int, db: Session = Depends(get_db)):
    """All requests submitted by a specific DC — shown on DC portal list page."""
    requests = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.dc_id == dc_id)
        .order_by(OperatorActivationRequest.submitted_at.desc())
        .all()
    )

    result = []
    for r in requests:
        remarks_history = [
            {
                "author_role": rm.author_role.upper(),
                "remark": rm.remark,
                "created_at": str(rm.created_at)[:16],
                "status_after": rm.status_after,
                "sender_username": rm.author.username if rm.author else "",
            }
            for rm in r.remarks
        ]

        # 🌟 UNIFORM SCHEMA FIX: Query district name dynamically using relationship attributes
        dist_name = r.district.district_name if r.district else "—"
        # 🌟 UNIFORM SCHEMA FIX: Normalize status to lowercase for accurate template matching
        clean_status = str(r.status or "sent_to_chips").strip().lower()

        result.append(
            {
                "id": r.id,
                "request_no": r.request_no if r.request_no else f"ACT-REQ-{r.id}",
                "operator_name": r.name_as_per_aadhaar,
                "operator_mobile": r.operator_mobile,
                "operator_aadhaar": r.operator_aadhaar,
                "operator_pan": r.pan_number,
                "district_name": dist_name,
                "status": clean_status,
                "submitted_at": str(r.submitted_at)[:16] if r.submitted_at else "",
                "reviewed_at": str(r.reviewed_at)[:16] if r.reviewed_at else None,
                "remarks_history": remarks_history,
            }
        )
    return result

# ─────────────────────────────────────────────
# CHIPS ADMIN ROUTES
# ─────────────────────────────────────────────

@router.get("/all")
def get_all_requests(db: Session = Depends(get_db)):
    """All requests across all DCs — shown on CHIPS admin dashboard."""
    requests = (
        db.query(OperatorActivationRequest)
        .order_by(OperatorActivationRequest.submitted_at.desc())
        .all()
    )

    result = []
    for r in requests:
        dist_name = r.district.district_name if r.district else "—"
        clean_status = str(r.status or "sent_to_chips").strip().lower()

        result.append(
            {
                "id": r.id,
                "request_no": r.request_no if r.request_no else f"ACT-REQ-{r.id}",
                "dc_id": r.dc_id,
                "district_id": r.district_id,
                "district_name": dist_name,
                "name_as_per_aadhaar": r.name_as_per_aadhaar,
                "operator_name": r.name_as_per_aadhaar,
                "operator_mobile": r.operator_mobile,
                "operator_aadhaar": r.operator_aadhaar,
                "operator_pan": r.pan_number,
                "status": clean_status,
                "remark_to_uidai": r.remark_to_uidai,
                "submitted_at": str(r.submitted_at)[:16] if r.submitted_at else "",
                "reviewed_at": str(r.reviewed_at)[:16] if r.reviewed_at else None,
                "reviewed_by": r.reviewed_by,
            }
        )
    return result



@router.get("/export-excel")
def export_to_excel(db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import openpyxl
    import io

    # 1. Query rows from database table having status as 'sent_to_uidai'
    requests_list = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.status == "sent_to_uidai")
        .order_by(OperatorActivationRequest.submitted_at.desc())
        .all()
    )

    # 2. Setup structural Workbook instance
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UIDAI Requests Archive"

    # 3. All database tracking fields mapping headers
    headers = [
        "ID",
        "Request Number",
        "DC ID",
        "District ID",
        "Role",
        "Name as per Aadhaar",
        "Registrar Code",
        "EA Code",
        "User Code",
        "NSEIT Certificate Number",
        "Operator Mobile",
        "Primary Email",
        "Operator Aadhaar",
        "NSEIT Certification Date",
        "Pincode",
        "Current Status",
        "Submitted At",
        "Reviewed At",
        "UIDAI Final Remarks",
    ]
    ws.append(headers)

    # 4. Pull and append ALL raw database records matching the state criteria
    for r in requests_list:
        ws.append(
            [
                r.id,
                r.request_no if r.request_no else "—",
                r.dc_id,
                r.district_id,
                r.role if r.role else "—",
                r.name_as_per_aadhaar,
                r.registrar_code if r.registrar_code else "—",
                r.ea_code if r.ea_code else "—",
                r.user_code if r.user_code else "—",
                r.nseit_certificate_number if r.nseit_certificate_number else "—",
                r.operator_mobile,
                r.primary_email if r.primary_email else "—",
                r.operator_aadhaar if r.operator_aadhaar else "—",
                (
                    str(r.nseit_certification_date)[:10]
                    if r.nseit_certification_date
                    else "—"
                ),
                (
                    str(r.nseit_certificate_expiry_date)[:10]
                    if r.nseit_certificate_expiry_date
                    else "—"
                ),
                r.pincode if r.pincode else "—",
                r.status,
                str(r.submitted_at)[:16] if r.submitted_at else "—",
                str(r.reviewed_at)[:16] if r.reviewed_at else "—",
                (
                    r.remark_to_uidai
                    if hasattr(r, "remark_to_uidai") and r.remark_to_uidai
                    else ""
                ),
            ]
        )

    # 5. Build raw stream buffer directly to prevent byte chunk loss
    stream = io.BytesIO()
    wb.save(stream)
    file_bytes = stream.getvalue()
    stream.close()

    # 6. Stream content natively with absolute spreadsheet content headers
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=sent_to_uidai_all_fields.xlsx",
            "Cache-Control": "no-cache",
        },
    )


@router.get("/export-excel/pending")
def export_pending_to_excel(ids: str = None, db: Session = Depends(get_db)):
    """Export Pending Operator Activation Queue to Excel.
    Optional ?ids=1,2,3 to export only specific (filtered) rows.
    """
    from fastapi.responses import StreamingResponse
    import openpyxl
    import io

    query = db.query(OperatorActivationRequest).filter(
        OperatorActivationRequest.status.in_(["sent_to_chips", "pending"])
    )
    if ids:
        id_list = [int(i.strip()) for i in ids.split(",") if i.strip().isdigit()]
        if id_list:
            query = query.filter(OperatorActivationRequest.id.in_(id_list))
    requests_list = query.order_by(OperatorActivationRequest.submitted_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Activation Queue"

    headers = [
        "S.No", "Request Number", "DC ID", "District ID", "Role",
        "Name as per Aadhaar", "Registrar Code", "EA Code", "User Code",
        "NSEIT Certificate Number", "Operator Mobile", "Primary Email",
        "Operator Aadhaar", "NSEIT Certification Date", "Pincode",
        "Current Status", "Submitted At",
    ]
    ws.append(headers)

    for idx, r in enumerate(requests_list, start=1):
        ws.append([
            idx,
            r.request_no if r.request_no else "—",
            r.dc_id,
            r.district_id,
            r.role if r.role else "—",
            r.name_as_per_aadhaar,
            r.registrar_code if r.registrar_code else "—",
            r.ea_code if r.ea_code else "—",
            r.user_code if r.user_code else "—",
            r.nseit_certificate_number if r.nseit_certificate_number else "—",
            r.operator_mobile,
            r.primary_email if r.primary_email else "—",
            r.operator_aadhaar if r.operator_aadhaar else "—",
            str(r.nseit_certification_date)[:10] if r.nseit_certification_date else "—",
            str(r.nseit_certificate_expiry_date)[:10] if r.nseit_certificate_expiry_date else "—",
            r.pincode if r.pincode else "—",
            r.status,
            str(r.submitted_at)[:16] if r.submitted_at else "—",
        ])

    stream = io.BytesIO()
    wb.save(stream)
    file_bytes = stream.getvalue()
    stream.close()

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=pending_activation_queue.xlsx",
            "Cache-Control": "no-cache",
        },
    )


@router.get("/export-excel/credentials")
def export_credentials_to_excel(ids: str = None, db: Session = Depends(get_db)):
    """Export Credentials Log History (approved / rejected / reverted) to Excel.
    Optional ?ids=1,2,3 to export only specific (filtered) rows.
    """
    from fastapi.responses import StreamingResponse
    import openpyxl
    import io

    query = db.query(OperatorActivationRequest).filter(
        OperatorActivationRequest.status.in_(["approved", "rejected", "reverted"])
    )
    if ids:
        id_list = [int(i.strip()) for i in ids.split(",") if i.strip().isdigit()]
        if id_list:
            query = query.filter(OperatorActivationRequest.id.in_(id_list))
    requests_list = query.order_by(OperatorActivationRequest.submitted_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials Log History"

    headers = [
        "S.No", "Request Number", "DC ID", "District ID", "Role",
        "Name as per Aadhaar", "Registrar Code", "EA Code", "User Code",
        "NSEIT Certificate Number", "Operator Mobile", "Primary Email",
        "Operator Aadhaar", "NSEIT Certificate Issue Date", "NSEIT Certificate Expiry Date",
        "Pincode", "Final Status", "Submitted At", "Reviewed At", "UIDAI / Admin Remarks",
    ]
    ws.append(headers)

    for idx, r in enumerate(requests_list, start=1):
        ws.append([
            idx,
            r.request_no if r.request_no else "—",
            r.dc_id,
            r.district_id,
            r.role if r.role else "—",
            r.name_as_per_aadhaar,
            r.registrar_code if r.registrar_code else "—",
            r.ea_code if r.ea_code else "—",
            r.user_code if r.user_code else "—",
            r.nseit_certificate_number if r.nseit_certificate_number else "—",
            r.operator_mobile,
            r.primary_email if r.primary_email else "—",
            r.operator_aadhaar if r.operator_aadhaar else "—",
            str(r.nseit_certification_date)[:10] if r.nseit_certification_date else "—",
            str(r.nseit_certificate_expiry_date)[:10] if r.nseit_certificate_expiry_date else "—",
            r.pincode if r.pincode else "—",
            r.status,
            str(r.submitted_at)[:16] if r.submitted_at else "—",
            str(r.reviewed_at)[:16] if r.reviewed_at else "—",
            r.remark_to_uidai if r.remark_to_uidai else "—",
        ])

    stream = io.BytesIO()
    wb.save(stream)
    file_bytes = stream.getvalue()
    stream.close()

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=credentials_log_history.xlsx",
            "Cache-Control": "no-cache",
        },
    )


@router.get("/{request_id}")
@router.get("/{request_id}/detail-json")
def get_request_detail(request_id: int, db: Session = Depends(get_db)):
    """Single request with all its documents — used on CHIPS detail/review page."""
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )

    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")

    docs = [
        {
            "doc_type": d.doc_type,
            "original_filename": d.original_filename,
            "file_path": d.file_path,
            "mime_type": d.mime_type,
            "uploaded_at": d.uploaded_at,
        }
        for d in r.documents
    ]

    remarks_history = [
        {
            "author_role": rm.author_role.upper(),
            "remark": rm.remark,
            "created_at": str(rm.created_at)[:16],
            "status_after": rm.status_after,
            "sender_username": rm.author.username if rm.author else "",
        }
        for rm in r.remarks
    ]

    dist_name = r.district.district_name if r.district else "—"
    clean_status = str(r.status or "sent_to_chips").strip().lower()
    latest_remark = r.remarks[-1].remark if r.remarks else None

    return {
        "id": r.id,
        "request_no": r.request_no,
        "dc_id": r.dc_id,
        "district_id": r.district_id,
        "district_name": dist_name,
        "operator_name": r.name_as_per_aadhaar,
        "operator_mobile": r.operator_mobile,
        "operator_aadhaar": r.operator_aadhaar,
        "operator_pan": r.pan_number,
        "primary_email": r.primary_email,
        "role": r.role,
        "registrar_code": r.registrar_code,
        "ea_code": r.ea_code,
        "user_code": r.user_code,
        "nseit_certificate_number": r.nseit_certificate_number,
        "nseit_certification_date": str(r.nseit_certification_date)[:10] if r.nseit_certification_date else None,
        "nseit_certificate_expiry_date": str(r.nseit_certificate_expiry_date)[:10] if r.nseit_certificate_expiry_date else None,
        "pincode": r.pincode,
        "status": clean_status,
        "rejection_reason": latest_remark,
        "chips_remarks": r.remark_to_uidai,
        "submitted_at": str(r.submitted_at)[:16] if r.submitted_at else None,
        "reviewed_at": str(r.reviewed_at)[:16] if r.reviewed_at else None,
        "reviewed_by": r.reviewed_by,
        "documents": docs,
        "remarks_history": remarks_history,
    }


@router.patch("/{request_id}/approve")
def approve_request(
    request_id: int,
    reviewed_by: int = Form(...),
    chips_remarks: str = Form(None),
    db: Session = Depends(get_db),
):
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )

    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")
    if r.status != "pending":
        raise HTTPException(status_code=400, detail=f"Request is already {r.status}.")

    r.status = "approved"
    r.reviewed_by = reviewed_by
    r.chips_remarks = chips_remarks
    r.reviewed_at = datetime.utcnow() + timedelta(hours=5, minutes=30)  # IST

    db.commit()
    return {"message": "Operator activated successfully.", "request_id": r.id}


@router.patch("/{request_id}/reject")
def reject_request(
    request_id: int,
    reviewed_by: int = Form(...),
    rejection_reason: str = Form(...),
    chips_remarks: str = Form(None),
    db: Session = Depends(get_db),
):
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )

    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")
    if r.status not in ["pending", "sent_to_chips", "sent_to_uidai"]:
        raise HTTPException(
            status_code=400, detail=f"Cannot revert a request with status: {r.status}"
        )

    r.status = "reverted"
    r.reviewed_by = reviewed_by
    r.chips_remarks = chips_remarks
    r.reviewed_at = datetime.utcnow() + timedelta(hours=5, minutes=30)  # IST

    # Create a remark record so DC can see the rejection reason
    remark = OperatorActivationRemark(
        request_id=r.id,
        author_id=reviewed_by,
        author_role="chips_admin",
        remark=rejection_reason,
        status_after="reverted",
    )
    db.add(remark)
    db.commit()
    return {
        "message": "Request reverted.",
        "request_id": r.id,
        "reason": rejection_reason,
    }


@router.patch("/{request_id}/send-to-uidai")
def send_to_uidai(
    request_id: int,
    reviewed_by: int = Form(...),
    uidai_remarks: str = Form(None),
    db: Session = Depends(get_db),
):
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )
    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")
    r.status = "sent_to_uidai"
    r.reviewed_by = reviewed_by
    r.remark_to_uidai = uidai_remarks
    r.reviewed_at = datetime.utcnow() + timedelta(hours=5, minutes=30)

    if uidai_remarks:
        remark = OperatorActivationRemark(
            request_id=r.id,
            author_id=reviewed_by,
            author_role="chips_admin",
            remark=f"Sent to UIDAI: {uidai_remarks}",
            status_after="sent_to_uidai",
        )
        db.add(remark)

    db.commit()
    return {"message": "Sent to UIDAI.", "request_id": r.id}


@router.patch("/{request_id}/uidai-approve")
def uidai_approve(
    request_id: int,
    reviewed_by: int = Form(...),
    uidai_remarks: str = Form(None),
    db: Session = Depends(get_db),
):
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )
    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")
    r.status = "approved"
    r.reviewed_by = reviewed_by
    r.remark_to_uidai = uidai_remarks
    r.reviewed_at = datetime.utcnow() + timedelta(hours=5, minutes=30)
    db.commit()
    return {"message": "Approved by UIDAI.", "request_id": r.id}


@router.patch("/{request_id}/uidai-reject")
def uidai_reject(
    request_id: int,
    reviewed_by: int = Form(...),
    uidai_remarks: str = Form(...),
    db: Session = Depends(get_db),
):
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )
    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")
    r.status = "rejected"
    r.reviewed_by = reviewed_by
    r.remark_to_uidai = uidai_remarks
    r.reviewed_at = datetime.utcnow() + timedelta(hours=5, minutes=30)

    remark = OperatorActivationRemark(
        request_id=r.id,
        author_id=reviewed_by,
        author_role="chips_admin",
        remark=f"Rejected by UIDAI: {uidai_remarks}",
        status_after="rejected",
    )
    db.add(remark)
    db.commit()
    return {"message": "Rejected by UIDAI.", "request_id": r.id}


@router.get("/{request_id}/detail")
def get_request_detail_full(request_id: int, db: Session = Depends(get_db)):
    # Simply points directly back to our synchronized payload schema to cut out duplicates
    return get_request_detail(request_id=request_id, db=db)

# Updated route to match the /dc/{id}/reapply URL design pattern
@router.post("/dc/{request_id}/reapply")
def reapply_request(
    request_id: int,
    dc_id: int = Form(...),
    operator_name: str = Form(None),
    operator_mobile: str = Form(...),
    operator_aadhaar: str = Form(None),
    operator_pan: str = Form(None),
    reapply_remark: str = Form(...),
    db: Session = Depends(get_db),
):
    r = (
        db.query(OperatorActivationRequest)
        .filter(OperatorActivationRequest.id == request_id)
        .first()
    )
    if not r:
        raise HTTPException(status_code=404, detail="Request not found.")

    # Only a reverted request can be corrected and sent back to CHIPS
    # 🌟 FIXED: Allows reapplying requests whether they were reverted by Admin or rejected by UIDAI
    if r.status not in ["reverted", "rejected"]:
        raise HTTPException(
            status_code=400, detail=f"Cannot reapply a request with status: {r.status}"
        )

    # Update operator details
    if operator_name:
        r.name_as_per_aadhaar = operator_name
    r.operator_mobile = operator_mobile
    r.operator_aadhaar = operator_aadhaar
    r.pan_number = operator_pan.upper() if operator_pan else r.pan_number

    # Reset status back to sent_to_chips so CHIPS admin receives the corrected request
    r.status = "sent_to_chips"
    r.reviewed_at = None

    # Save DC remark to the conversation history model tracking table
    remark = OperatorActivationRemark(
        request_id=r.id,
        author_id=dc_id,
        author_role="dc",
        remark=reapply_remark,
        status_after="sent_to_chips",
    )
    db.add(remark)
    db.commit()

    return {"message": "Request reapplied successfully.", "request_id": r.id}


# ─────────────────────────────────────────────
# FILE SERVE ENDPOINT
# ─────────────────────────────────────────────

@router.get("/{request_id}/file/{doc_type}")
def serve_document(request_id: int, doc_type: str, db: Session = Depends(get_db)):
    """Stream a stored document file to the browser (inline view)."""
    from fastapi.responses import FileResponse

    if doc_type not in VALID_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Invalid document type.")

    doc = (
        db.query(ActivationDocument)
        .filter(
            ActivationDocument.request_id == request_id,
            ActivationDocument.doc_type == doc_type,
        )
        .first()
    )

    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")

    if not os.path.exists(doc.file_path):
        raise HTTPException(status_code=404, detail="File missing on disk.")

    return FileResponse(
        path=doc.file_path,
        media_type=doc.mime_type or "application/octet-stream",
        filename=doc.original_filename,
        headers={"Content-Disposition": f"inline; filename=\"{doc.original_filename}\""},
    )
